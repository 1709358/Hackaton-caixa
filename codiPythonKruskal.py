import openpyxl 
import sys 
from itertools import pairwise

matrix = open("/home/andreu/universitat/hackAton/ferKruskal/matriuRuta1.txt")
matrixInteger = []
for i in matrix:
    listToset = []
    for j in i.split(sep=' '):
        listToset.append(int(j))
    matrixInteger.append(listToset)

matrixValues = []
begin = 77
end = 106
lenghData = end - begin

for i in range(lenghData):
    lista = []
    for j in range(lenghData):
        lista.append(matrixInteger[i][j])
    matrixValues.append(lista)
posicionVisitades = []

wb = openpyxl.load_workbook('/home/andreu/universitat/hackAton/ferKruskal/Datos.xlsx')
ws = wb['Hoja1']
costesEsrtar = [ws.cell(row=i,column=17).value for i in range(begin,end+1)]
nomLlocs = [ws.cell(row=i,column=13).value for i in range(begin,end+1)]

cami=[]

    
    
pivot = 0
posicionVisitades.append(0)
listaPivots = [pivot]
camiOrigen=[]
camiDesti=[]
costCami = []
for i in range(int(lenghData/2)):
    cost_minim=sys.maxsize 
    pos_minim_origen = sys.maxsize 
    pos_minim_desti = sys.maxsize 
    for pivotElement,posPivot in zip(matrixValues,range(len(matrixValues))):
        for element,pos in zip(pivotElement,range(len(matrixValues))):
            if ((posPivot not in camiDesti) and (pos not in camiDesti)):
                if((pos not in camiOrigen) and (posPivot not in camiOrigen)):
                    if (cost_minim > element + costesEsrtar[pos]) and element!=0:
                        cost_minim = element + costesEsrtar[pos]
                        pos_minim_desti = pos
                        pivot = posPivot

    
    costCami.append(element + costesEsrtar[pos])
    listaPivots.append(pos_minim_desti)
    camiDesti.append(pivot)
    camiOrigen.append(pos_minim_desti)

    cami.append([pivot,pos_minim_desti])


value = 0
for i in costCami:
    value+=i


novaLlista = []
for i,j in zip(camiOrigen,camiDesti):
    novaLlista.append(i)
    novaLlista.append(j)

nouValueMinim = 0
for i in pairwise(novaLlista):
    nouValueMinim+=matrixValues[i[0]][i[1]]+costesEsrtar[i[1]]


llistadeLlocOrdenada = []
for i in novaLlista:
    llistadeLlocOrdenada.append(nomLlocs[i])
print("Llista de dades ordenada desde origen fins desti")
print(llistadeLlocOrdenada)
print("Cost trayecta: " + str(nouValueMinim))

#print(value)


#si agafas albiol de l'area 4 redueixes en 2h
